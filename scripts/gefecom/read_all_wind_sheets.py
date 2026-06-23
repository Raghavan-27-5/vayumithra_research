import openpyxl

wb = openpyxl.load_workbook(r"C:\Projects\raghavan\GEFCom2014\GEFCom2014 Data\Provisional_Leaderboard_V2.xlsx")

for sn in ["W-score-0", "W-score-1", "W-score-2", "W-score-3", "W-score", "W-eligible", "W-log"]:
    ws = wb[sn]
    print(f"\n{'='*100}")
    print(f"{sn}  (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*100}")
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 20))]
        print(vals)
