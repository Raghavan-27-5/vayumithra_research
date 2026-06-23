"""Update Excel report with GEFCom2014 leaderboard comparison."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np

# ── Load saved results ─────────────────────────────────────────────────────
data = np.load(r"C:\Projects\raghavan\vayumithra_research\results\gefecom_task15_results.npz")
model_preds = data["model_preds"]
bench_preds = data["bench_preds"]
actuals = data["actuals"]
quantiles = data["quantiles"]

def pinball(y, yh, q):
    e = y - yh
    return float(np.mean(np.maximum(q*e, (q-1)*e)))

per_q_model = np.array([pinball(actuals, model_preds[:, qi], q) for qi, q in enumerate(quantiles)])
per_q_bench = np.array([pinball(actuals, bench_preds[:, qi], q) for qi, q in enumerate(quantiles)])

# ── Load leaderboard ───────────────────────────────────────────────────────
wb_lb = openpyxl.load_workbook(r"C:\Projects\raghavan\GEFCom2014\GEFCom2014 Data\Provisional_Leaderboard_V2.xlsx")
ws_lb = wb_lb["W-score-0"]
leaderboard = {}
bench_name = None
for r in range(2, ws_lb.max_row + 1):
    name = ws_lb.cell(r, 1).value
    if name is None: continue
    scores = [ws_lb.cell(r, c).value for c in range(2, 14)]
    scores = [s for s in scores if isinstance(s, (int, float))]
    if not scores: continue
    if "Benchmark" in str(name):
        bench_name = name
    leaderboard[name] = np.mean(scores)

bench_lb = leaderboard.pop(bench_name, 0)

# ── Open existing report ───────────────────────────────────────────────────
src = r"C:\Projects\raghavan\vayumithra_research\results\probabilistic_forecasting_report.xlsx"
out = r"C:\Users\Nandha\AppData\Local\Temp\opencode\probabilistic_forecasting_report.xlsx"
wb = openpyxl.load_workbook(src)

hdr_font = Font(bold=True, size=12)
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_font_white = Font(bold=True, size=11, color="FFFFFF")
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Remove old comparison sheet if exists ──────────────────────────────────
for sn in ["GEFCom_Comparison", "Leaderboard"]:
    if sn in wb.sheetnames:
        del wb[sn]

ws_cmp = wb.create_sheet("GEFCom_Comparison")
col = 1
row = 1

# Helper
def write_row(ws, r, data, font=None, fill=None):
    for c, v in enumerate(data, 1):
        cell = ws.cell(r, c, v)
        if font: cell.font = font
        if fill: cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

# ═══════════════════════════════════════════════════════════════════════════
# TABLE 1: Our Model (iTransformer) vs Benchmark  Per-Quantile
# ═══════════════════════════════════════════════════════════════════════════
write_row(ws_cmp, row, ["GEFCom2014 Task 15  Per-Quantile Pinball: iTransformer vs Benchmark",
                        "", "", "", ""], font=Font(bold=True, size=13))
row += 2
write_row(ws_cmp, row, ["Quantile", "Model Pinball", "Bench Pinball", "M/B Ratio", "Interpretation"],
          font=hdr_font_white, fill=hdr_fill)
row += 1

# Selected quantiles
show_q = [0, 9, 24, 49, 74, 89, 98]  # P1, P10, P25, P50, P75, P90, P99
for qi in show_q:
    q = quantiles[qi]
    mp = per_q_model[qi]
    bp = per_q_bench[qi]
    ratio = mp / bp if bp > 0 else float("inf")
    interp = "Poor" if ratio > 3 else ("Moderate" if ratio > 1.5 else "Good")
    write_row(ws_cmp, row, [f"P{int(q*100)} (q={q:.2f})",
                            f"{mp:.6f}", f"{bp:.6f}",
                            f"{ratio:.2f}x", interp])
    row += 1

# Average row
avg_m = float(np.mean(per_q_model))
avg_b = float(np.mean(per_q_bench))
write_row(ws_cmp, row, ["AVERAGE", f"{avg_m:.6f}", f"{avg_b:.6f}",
                        f"{avg_m/avg_b:.2f}x", "Benchmark is better"],
          font=Font(bold=True))
row += 1
# Our ratio to top
write_row(ws_cmp, row, ["kPower (1st place)", "", "", "0.43x bench", "Winning level"])
row += 2

# ═══════════════════════════════════════════════════════════════════════════
# TABLE 2: Leaderboard Comparison
# ═══════════════════════════════════════════════════════════════════════════
write_row(ws_cmp, row, ["GEFCom2014 Wind Leaderboard  Avg Pinball per Task (Tasks 1-12)",
                        "", "", "", ""], font=Font(bold=True, size=13))
row += 2
write_row(ws_cmp, row, ["Rank", "Team", "Avg Pinball", "vs Benchmark", "Note"],
          font=hdr_font_white, fill=hdr_fill)
row += 1

sorted_lb = sorted(leaderboard.items(), key=lambda x: x[1])
for rank, (name, avg) in enumerate(sorted_lb, 1):
    ratio = avg / bench_lb if bench_lb else 1
    write_row(ws_cmp, row, [rank, name, f"{avg:.6f}", f"{ratio:.2f}x bench", ""])
    row += 1

# Benchmark row
write_row(ws_cmp, row, ["-", "Benchmark", f"{bench_lb:.6f}", "1.00x", "Organizer baseline"],
          font=Font(bold=True))
row += 1

# Our model comparison
our_vs_bench = avg_m / avg_b
our_vs_kpower = (avg_m / avg_b) / 0.43
write_row(ws_cmp, row, ["-", "Our iTransformer (Task 15)", f"{avg_m:.4f}",
                        f"{our_vs_bench:.2f}x bench",
                        f"{our_vs_kpower:.1f}x worse than kPower"],
          font=Font(bold=True, color="FF0000"))
row += 2

# ═══════════════════════════════════════════════════════════════════════════
# TABLE 3: Per-Zone Comparison (Model vs Benchmark)
# ═══════════════════════════════════════════════════════════════════════════
write_row(ws_cmp, row, ["Per-Zone Breakdown  Task 15", "", "", "", ""],
          font=Font(bold=True, size=13))
row += 2
write_row(ws_cmp, row, ["Zone", "iTransformer Pinball", "Benchmark Pinball",
                        "Ratio", "iTransformer P50 MAE"],
          font=hdr_font_white, fill=hdr_fill)
row += 1

N_ZONES = 10
for z in range(1, N_ZONES + 1):
    zdata = data.get(f"zone_{z}") if hasattr(data, "get") else None

# Recompute per-zone from saved data (we didnt save per-zone separately)
# Recompute from the raw data
all_actuals_flat = actuals
all_preds_flat = model_preds
all_bench_flat = bench_preds

# Actually our npz doesnt have per-zone labels, but we can just note it
write_row(ws_cmp, row, ["All 10 zones", f"{avg_m:.6f}", f"{avg_b:.6f}",
                        f"{avg_m/avg_b:.2f}x",
                        f"{float(np.mean(np.abs(all_actuals_flat - all_preds_flat[:, 49]))):.4f}"])
row += 1

# ═══════════════════════════════════════════════════════════════════════════
# Column widths
# ═══════════════════════════════════════════════════════════════════════════
ws_cmp.column_dimensions["A"].width = 28
ws_cmp.column_dimensions["B"].width = 22
ws_cmp.column_dimensions["C"].width = 22
ws_cmp.column_dimensions["D"].width = 16
ws_cmp.column_dimensions["E"].width = 30

# ── Update Summary sheet ───────────────────────────────────────────────────
ws_sum = wb["Summary"]
# Find next empty row
sr = ws_sum.max_row + 2
ws_sum.cell(sr, 1, "GEFCom Leaderboard Comparison").font = Font(bold=True, size=11)
sr += 1
ws_sum.cell(sr, 1, "Our avg pinball")
ws_sum.cell(sr, 2, f"{avg_m:.4f}")
sr += 1
ws_sum.cell(sr, 1, "Benchmark avg pinball")
ws_sum.cell(sr, 2, f"{avg_b:.4f}")
sr += 1
ws_sum.cell(sr, 1, "Our ratio to benchmark")
ws_sum.cell(sr, 2, f"{avg_m/avg_b:.2f}x")
sr += 1
ws_sum.cell(sr, 1, "kPower ratio to benchmark")
ws_sum.cell(sr, 2, "0.43x")
sr += 1
ws_sum.cell(sr, 1, "Gap to kPower")
ws_sum.cell(sr, 2, f"{our_vs_kpower:.1f}x worse")
sr += 1
ws_sum.cell(sr, 1, "Root cause")
ws_sum.cell(sr, 2, "Explanatory variables (U10/V10/U100/V100 future forecasts) not used")

wb.save(out)
print(f"Updated: {out}")
print(f"Added sheet: GEFCom_Comparison")
print(f"Updated: Summary")
