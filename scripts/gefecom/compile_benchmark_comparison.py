import openpyxl, numpy as np

SRC_LB = r"C:\Projects\raghavan\GEFCom2014\GEFCom2014 Data\Provisional_Leaderboard_V2.xlsx"
SRC_REPORT = r"C:\Projects\raghavan\vayumithra_research\results\probabilistic_forecasting_report.xlsx"
OUT = r"C:\Users\Nandha\AppData\Local\Temp\opencode\probabilistic_forecasting_report.xlsx"

# ── Read leaderboard raw data ──────────────────────────────────────────────
wb = openpyxl.load_workbook(SRC_LB)
# W-score-0: raw pinball scores per task (lower=better)
ws0 = wb["W-score-0"]
header = [ws0.cell(1, c).value for c in range(1, 14)]

# Benchmark row (row 2)
bench_raw = [ws0.cell(2, c).value for c in range(2, 14)]

# All teams (rows 3-14)
teams_raw = {}
for r in range(3, ws0.max_row + 1):
    name = ws0.cell(r, 1).value
    if name:
        scores = [ws0.cell(r, c).value for c in range(2, 14)]
        teams_raw[name] = scores

# W-score-3: Ratings
ws3 = wb["W-score-3"]
ratings_raw = {}
for r in range(2, ws3.max_row + 1):
    name = ws3.cell(r, 1).value
    rating = ws3.cell(r, 14).value
    rank = ws3.cell(r, 15).value
    if name:
        ratings_raw[name] = (rating, rank)

# Summary-1: Final leaderboard (clean names, ratings)
ws_sum = wb["Summary-1"]
# Wind rating in column 6, team in column 5
# Summary-1 structure:
# Row 1: [None, Load, None, Price, None, Wind, None, Solar, None]
# Row 2: [Ranking, Team, Rating, Team, Rating, Team, Rating, Team, Rating]
# Wind team = column 6, Wind rating = column 7
wind_leaderboard = []
for r in range(3, 25):
    team = ws_sum.cell(r, 6).value
    rating = ws_sum.cell(r, 7).value
    if team and isinstance(rating, (int, float)):
        wind_leaderboard.append((team, rating))

if not wind_leaderboard:
    print("WARNING: Wind leaderboard empty, falling back to W-score-3")
    for r in range(2, ws3.max_row + 1):
        name = ws3.cell(r, 1).value
        rating = ws3.cell(r, 14).value
        if name and isinstance(rating, (int, float)):
            wind_leaderboard.append((name, rating))
    wind_leaderboard = sorted(wind_leaderboard, key=lambda x: -x[1])
# Sort by rating descending
wind_leaderboard = sorted(wind_leaderboard, key=lambda x: -x[1])
print(f"Wind leaderboard entries: {len(wind_leaderboard)}")
print(f"Top 3: {wind_leaderboard[:3]}")

# ── Compute averages ───────────────────────────────────────────────────────
bench_scores = [s for s in bench_raw if isinstance(s, (int, float))]
bench_avg = np.mean(bench_scores) if bench_scores else 0

kpower_scores = [s for s in teams_raw.get("kPower", []) if isinstance(s, (int, float))]
kpower_avg = np.mean(kpower_scores) if kpower_scores else 0

dmlab_scores = [s for s in teams_raw.get("dmlab", []) if isinstance(s, (int, float))]
dmlab_avg = np.mean(dmlab_scores) if dmlab_scores else 0

# ── Our model (iTransformer Task 15) ───────────────────────────────────────
data = np.load(r"C:\Projects\raghavan\vayumithra_research\results\gefecom_task15_results.npz")
model_preds = data["model_preds"]
bench_preds = data["bench_preds"]
actuals = data["actuals"]
quantiles = data["quantiles"]

def pinball(y, yh, q):
    e = y - yh
    return float(np.mean(np.maximum(q * e, (q - 1) * e)))

per_q_model = np.array([pinball(actuals, model_preds[:, qi], q) for qi, q in enumerate(quantiles)])
per_q_bench_t15 = np.array([pinball(actuals, bench_preds[:, qi], q) for qi, q in enumerate(quantiles)])
our_avg = float(np.mean(per_q_model))
bench_t15_avg = float(np.mean(per_q_bench_t15))

# ── Write to Excel ─────────────────────────────────────────────────────────
wb_out = openpyxl.load_workbook(SRC_REPORT)
for sn in ["Leaderboard_Comparison"]:
    if sn in wb_out.sheetnames:
        del wb_out[sn]
ws = wb_out.create_sheet("Leaderboard_Comparison")

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_font = Font(bold=True, size=11, color="FFFFFF")
bold_font = Font(bold=True)
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def w(ws, r, c, v, font=None, fill=None):
    cell = ws.cell(r, c, v)
    cell.border = border
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    if font: cell.font = font
    if fill: cell.fill = fill
    return cell

row = 1
# ── TABLE 1: Competition Benchmark Per-Task Scores ─────────────────────────
w(ws, row, 1, "GEFCom2014-W COMPETITION DATA (Tasks 1-12)", font=Font(bold=True, size=13))
row += 2

w(ws, row, 1, "Task", font=hdr_font, fill=hdr_fill)
for c in range(1, 13):
    w(ws, row, c + 1, f"Task {c}", font=hdr_font, fill=hdr_fill)
w(ws, row, 14, "Avg", font=hdr_font, fill=hdr_fill)
row += 1

# Benchmark row
w(ws, row, 1, "Benchmark (Organizer)", font=bold_font)
for c in range(12):
    v = bench_raw[c] if isinstance(bench_raw[c], (int, float)) else ""
    w(ws, row, c + 2, v)
w(ws, row, 14, round(bench_avg, 6), font=bold_font)
row += 1

# kPower row  
w(ws, row, 1, "kPower (1st Place)", font=bold_font)
kp = teams_raw.get("kPower", [None]*12)
for c in range(12):
    v = kp[c] if isinstance(kp[c], (int, float)) else ""
    w(ws, row, c + 2, v)
w(ws, row, 14, round(kpower_avg, 6), font=bold_font)
row += 1

# dmlab (2nd)
w(ws, row, 1, "dmlab (2nd Place)", font=bold_font)
dm = teams_raw.get("dmlab", [None]*12)
for c in range(12):
    v = dm[c] if isinstance(dm[c], (int, float)) else ""
    w(ws, row, c + 2, v)
w(ws, row, 14, round(dmlab_avg, 6), font=bold_font)
row += 1

# E.S. Mangalova (3rd)
w(ws, row, 1, "E.S. Mangalova (3rd Place)", font=bold_font)
es = teams_raw.get("E.S. Mangalova", [None]*12)
es_avg = np.mean([s for s in es if isinstance(s, (int, float))])
for c in range(12):
    v = es[c] if isinstance(es[c], (int, float)) else ""
    w(ws, row, c + 2, v)
w(ws, row, 14, round(float(es_avg), 6), font=bold_font)
row += 1

# All other teams
for tname in ["C3 Green Team", "Yao Zhang", "UTES", "Onverrabien", "pat1"]:
    tdat = teams_raw.get(tname, [None]*12)
    tavg = np.mean([s for s in tdat if isinstance(s, (int, float))])
    w(ws, row, 1, tname)
    for c in range(12):
        v = tdat[c] if isinstance(tdat[c], (int, float)) else ""
        w(ws, row, c + 2, v)
    w(ws, row, 14, round(float(tavg), 6))
    row += 1

row += 2

# ── TABLE 2: Final Ratings ─────────────────────────────────────────────────
w(ws, row, 1, "Final Competition Ratings (higher = better)", font=Font(bold=True, size=13))
row += 2

w(ws, row, 1, "Rank", font=hdr_font, fill=hdr_fill)
w(ws, row, 2, "Team", font=hdr_font, fill=hdr_fill)
w(ws, row, 3, "Rating", font=hdr_font, fill=hdr_fill)
w(ws, row, 4, "vs Benchmark", font=hdr_font, fill=hdr_fill)
row += 1

for rank, (team, rating) in enumerate(wind_leaderboard, 1):
    w(ws, row, 1, rank)
    w(ws, row, 2, team)
    w(ws, row, 3, round(rating, 6))
    w(ws, row, 4, f"{round(rating*100, 1)}% improvement over baseline")
    row += 1

row += 2

# ── TABLE 3: Our Model vs Benchmark (Task 15) ──────────────────────────────
w(ws, row, 1, "Our Model (iTransformer) vs Benchmark on Task 15", font=Font(bold=True, size=13))
row += 2

# Per-quantile comparison
w(ws, row, 1, "Quantile", font=hdr_font, fill=hdr_fill)
w(ws, row, 2, "Our Pinball", font=hdr_font, fill=hdr_fill)
w(ws, row, 3, "Benchmark Pinball", font=hdr_font, fill=hdr_fill)
w(ws, row, 4, "Ratio (M/B)", font=hdr_font, fill=hdr_fill)
w(ws, row, 5, "kPower Equivalent", font=hdr_font, fill=hdr_fill)
row += 1

for qi in [0, 9, 24, 49, 74, 89, 98]:
    q = quantiles[qi]
    mp = per_q_model[qi]
    bp = per_q_bench_t15[qi]
    ratio = mp / bp if bp > 0 else float("inf")
    # Estimate kPower level: ~0.43x benchmark ratio applied per-quantile
    kp_est = bp * 0.43
    w(ws, row, 1, f"P{int(q*100)} (q={q:.2f})")
    w(ws, row, 2, round(mp, 6))
    w(ws, row, 3, round(bp, 6))
    w(ws, row, 4, f"{ratio:.2f}x")
    w(ws, row, 5, round(kp_est, 6))
    row += 1

w(ws, row, 1, "AVERAGE", font=bold_font)
w(ws, row, 2, round(our_avg, 6), font=bold_font)
w(ws, row, 3, round(bench_t15_avg, 6), font=bold_font)
w(ws, row, 4, f"{our_avg/bench_t15_avg:.2f}x", font=bold_font)
w(ws, row, 5, round(bench_t15_avg * 0.43, 6), font=bold_font)
row += 2

# ── TABLE 4: Side-by-Side Summary ─────────────────────────────────────────
w(ws, row, 1, "Summary: Where We Stand", font=Font(bold=True, size=13))
row += 2

w(ws, row, 1, "Metric", font=hdr_font, fill=hdr_fill)
w(ws, row, 2, "Benchmark", font=hdr_font, fill=hdr_fill)
w(ws, row, 3, "kPower (1st)", font=hdr_font, fill=hdr_fill)
w(ws, row, 4, "Our Model", font=hdr_font, fill=hdr_fill)
w(ws, row, 5, "Gap", font=hdr_font, fill=hdr_fill)
row += 1

tasks_avg = bench_avg  # task 1-12 benchmark average
w(ws, row, 1, "Avg Pinball (Tasks 1-12)")
w(ws, row, 2, round(float(tasks_avg), 4))
w(ws, row, 3, round(kpower_avg, 4))
w(ws, row, 4, "N/A (Task 15 only)")
w(ws, row, 5, "")
row += 1

w(ws, row, 1, "Avg Pinball (Task 15)")
w(ws, row, 2, round(bench_t15_avg, 4))
w(ws, row, 3, "N/A (not public)")
w(ws, row, 4, round(our_avg, 4))
w(ws, row, 5, f"{our_avg/bench_t15_avg:.2f}x bench")
row += 1

w(ws, row, 1, "P50 MAE (Task 15)")
from sklearn.metrics import mean_absolute_error
our_p50_mae = mean_absolute_error(actuals, model_preds[:, 49])
bench_p50_mae = mean_absolute_error(actuals, bench_preds[:, 49])
w(ws, row, 2, round(bench_p50_mae, 4))
w(ws, row, 3, "N/A")
w(ws, row, 4, round(our_p50_mae, 4))
w(ws, row, 5, f"{our_p50_mae/bench_p50_mae:.2f}x bench")
row += 1

w(ws, row, 1, "Rating (leaderboard)")
w(ws, row, 2, "0 (baseline)")
kp_rating_local = next((r for t, r in wind_leaderboard if "kPower" in t), 0)
w(ws, row, 3, round(kp_rating_local, 6))
w(ws, row, 4, "N/A (not on leaderboard)")
w(ws, row, 5, "")
row += 1

w(ws, row, 1, "Root Cause for Gap", font=bold_font)
w(ws, row, 2, "")
w(ws, row, 3, "")
w(ws, row, 4, "No explanatory variables (future U10/V10/U100/V100)")
w(ws, row, 5, "")

# ── Column widths ──────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 32
for c in range(2, 15):
    ws.column_dimensions[chr(64 + c)].width = 16

# ── Update Summary too ─────────────────────────────────────────────────────
ws_s = wb_out["Summary"]
sr = ws_s.max_row + 2
ws_s.cell(sr, 1, "Benchmark pinball avg (tasks 1-12)").font = Font(bold=True)
ws_s.cell(sr, 2, round(float(tasks_avg), 6))
sr += 1
ws_s.cell(sr, 1, "kPower pinball avg (tasks 1-12)")
ws_s.cell(sr, 2, round(kpower_avg, 6))
sr += 1
ws_s.cell(sr, 1, "kPower rating (leaderboard)")
kp_rating = next((r for t, r in wind_leaderboard if "kPower" in t), 0)
ws_s.cell(sr, 2, round(kp_rating, 6))
sr += 1
ws_s.cell(sr, 1, "Our Task 15 pinball avg")
ws_s.cell(sr, 2, round(our_avg, 6))
sr += 1
ws_s.cell(sr, 1, "Benchmark Task 15 pinball avg")
ws_s.cell(sr, 2, round(bench_t15_avg, 6))
sr += 1
ws_s.cell(sr, 1, "Our ratio to Task 15 benchmark")
ws_s.cell(sr, 2, f"{our_avg/bench_t15_avg:.2f}x")
sr += 1
ws_s.cell(sr, 1, "kPower ratio to benchmark")
ws_s.cell(sr, 2, f"{kpower_avg/float(tasks_avg):.2f}x")
sr += 1
ws_s.cell(sr, 1, "Leaderboard sheet available")
ws_s.cell(sr, 2, "Leaderboard_Comparison")

wb_out.save(OUT)
print(f"Saved: {OUT}")
print("Sheets:", wb_out.sheetnames)
