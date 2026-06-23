"""
Generate Excel Report for iTransformer+NHiTS Probabilistic Forecasting.

Sheets:
  1. Summary          — project overview, datasets, key findings
  2. VayuMithra       — VayuMithra 10-station results (H1..H24)
  3. GEFCom2014       — GEFCom Task 15 benchmark comparison
  4. Architecture     — model architecture diagram and description
  5. Loss_Functions   — pinball loss, monotonicity, Winkler score
"""
import json, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Style helpers ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
BODY_FONT = Font(name="Calibri", size=10)
BOLD = Font(name="Calibri", bold=True, size=10)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
BAD_FILL = PatternFill("solid", fgColor="FCE4EC")

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def style_data_cell(cell, bold=False):
    cell.font = BOLD if bold else BODY_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center")

def write_row(ws, row, values, bold=False):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        style_data_cell(cell, bold)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: Summary
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Summary"
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 80

ws.cell(row=1, column=1, value="iTransformer+NHiTS Probabilistic Forecasting").font = TITLE_FONT
ws.merge_cells("A1:B1")
ws.cell(row=2, column=1, value="").font = BODY_FONT

summary_data = [
    ("Project", "Probabilistic wind speed / wind power forecasting"),
    ("Model", "iTransformer backbone + Per-Variate Projection + QuantileProjectionHead"),
    ("Quantiles", "P10, P50, P90, P99 (VayuMithra); 99 quantiles 0.01..0.99 (GEFCom)"),
    ("Loss Function", "Pinball Loss (Quantile Loss) with monotonicity enforcement"),
    ("Datasets", "VayuMithra (10-station uv, 2013-2022) & GEFCom2014-W (10-zone, 2012-2013)"),
    ("",
     "For VayuMithra: predicting wind speed at station s5 from 50 variates."),
    ("",
     "For GECCom2014: predicting normalized wind power (0-1) for 10 zones from 50 features."),
    ("Training Config", "d_model=128, e_layers=2, n_heads=4, d_ff=512, dropout=0.15, lr=3e-4"),
    ("Walk-forward CV", "VayuMithra: 4 temporal folds (2019/2020/2021/2022). GEFCom: single split."),
    ("",
     ""),
    ("Key Finding 1", "H1 wind speed coverage well-calibrated (0.822 vs target 0.80)"),
    ("Key Finding 2", "P50 MAE degrades with horizon: 0.23 (H1) → 1.74 (H24)"),
    ("Key Finding 3", "GEFCom benchmark outperforms pure time-series model (needs exp vars)"),
    ("",
     ""),
    ("Files", ""),
    ("Model", "vayumithra_research/src/models/probabilistic_model.py"),
    ("Training Script", "vayumithra_research/scripts/train_probabilistic.py"),
    ("Evaluation Script", "vayumithra_research/scripts/evaluate_probabilistic.py"),
    ("GEFCom Pipeline", "scripts/gefecom/run_gefecom.py"),
    ("Checkpoints", "vayumithra_research/results/models/probabilistic/*.pt"),
]

for i, (k, v) in enumerate(summary_data, 4):
    ws.cell(row=i, column=1, value=k).font = BOLD if k and not k.startswith("Key") else BODY_FONT
    ws.cell(row=i, column=2, value=v).font = BODY_FONT
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=2).border = BORDER

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: VayuMithra
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("VayuMithra")
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 14
ws2.column_dimensions["H"].width = 14
ws2.column_dimensions["I"].width = 14
ws2.column_dimensions["J"].width = 12

ws2.cell(row=1, column=1, value="VayuMithra 10-Station Probabilistic Forecasting Results").font = TITLE_FONT
ws2.merge_cells("A1:J1")

ws2.cell(row=3, column=1, value="Fold 1 — Test period 2020 (station s5 wind speed)").font = SECTION_FONT
ws2.merge_cells("A3:J3")

headers = ["Horizon", "P50 MAE", "P50 RMSE", "Coverage\nP10-P90", "Width\n(m/s)",
           "Winkler", "Pinball\nP10", "Pinball\nP50", "Pinball\nP90", "Pinball\nAvg"]
for i, h in enumerate(headers, 1):
    ws2.cell(row=5, column=i, value=h)
style_header_row(ws2, 5, len(headers))

# VayuMithra results
vm_results = {
    1: {"p50_mae": 0.2298, "p50_rmse": 0.3098, "coverage_p10_p90": 0.8219, "width_p10_p90": 0.80,
        "winkler_p10_p90": 1.1417, "pinball_p10": 0.0290, "pinball_p50": 0.1149, "pinball_p90": 0.0348, "pinball_avg": 0.0597},
    2: {"p50_mae": 0.4478, "p50_rmse": 0.5883, "coverage_p10_p90": 0.7390, "width_p10_p90": 1.27,
        "winkler_p10_p90": 2.1129, "pinball_p10": 0.0574, "pinball_p50": 0.2239, "pinball_p90": 0.0585, "pinball_avg": 0.1136},
    3: {"p50_mae": 0.6579, "p50_rmse": 0.8590, "coverage_p10_p90": 0.6743, "width_p10_p90": 1.63,
        "winkler_p10_p90": 3.2260, "pinball_p10": 0.0896, "pinball_p50": 0.3290, "pinball_p90": 0.0809, "pinball_avg": 0.1719},
    4: {"p50_mae": 1.0189, "p50_rmse": 1.3290, "coverage_p10_p90": 0.7049, "width_p10_p90": 2.70,
        "winkler_p10_p90": 4.8508, "pinball_p10": 0.1333, "pinball_p50": 0.5095, "pinball_p90": 0.1234, "pinball_avg": 0.2595},
    5: {"p50_mae": 1.2053, "p50_rmse": 1.5736, "coverage_p10_p90": 0.7453, "width_p10_p90": 3.57,
        "winkler_p10_p90": 5.8152, "pinball_p10": 0.1570, "pinball_p50": 0.6027, "pinball_p90": 0.1669, "pinball_avg": 0.3106},
    6: {"p50_mae": 1.1953, "p50_rmse": 1.5563, "coverage_p10_p90": 0.7037, "width_p10_p90": 3.13,
        "winkler_p10_p90": 5.6585, "pinball_p10": 0.1461, "pinball_p50": 0.5977, "pinball_p90": 0.1531, "pinball_avg": 0.3029},
    12: {"p50_mae": 1.5532, "p50_rmse": 2.0420, "coverage_p10_p90": 0.6616, "width_p10_p90": 3.69,
         "winkler_p10_p90": 7.6331, "pinball_p10": 0.2001, "pinball_p50": 0.7766, "pinball_p90": 0.2093, "pinball_avg": 0.4034},
    24: {"p50_mae": 1.7371, "p50_rmse": 2.2893, "coverage_p10_p90": 0.7224, "width_p10_p90": 4.78,
         "winkler_p10_p90": 8.3408, "pinball_p10": 0.2113, "pinball_p50": 0.8686, "pinball_p90": 0.2360, "pinball_avg": 0.4458},
}

for row_idx, h in enumerate([1, 2, 3, 4, 5, 6, 12, 24], 6):
    r = vm_results[h]
    vals = [h, r["p50_mae"], r["p50_rmse"], r["coverage_p10_p90"], r["width_p10_p90"],
            r["winkler_p10_p90"], r["pinball_p10"], r["pinball_p50"], r["pinball_p90"], r["pinball_avg"]]
    write_row(ws2, row_idx, vals)
    # Color coverage cells: green if >= 0.75, red if < 0.65
    cov_cell = ws2.cell(row=row_idx, column=4)
    if r["coverage_p10_p90"] >= 0.75:
        cov_cell.fill = GOOD_FILL
    elif r["coverage_p10_p90"] < 0.65:
        cov_cell.fill = BAD_FILL

ws2.cell(row=15, column=1, value="Deterministic iTransformer+NHiTS comparison (H1): P50 MAE target ≈ 0.12 m/s").font = BODY_FONT
ws2.merge_cells("A15:J15")
ws2.cell(row=16, column=1, value="Model underperforms deterministic by ~2x due to harder pinball optimization").font = BODY_FONT
ws2.merge_cells("A16:J16")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: GEFCom2014
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("GEFCom2014")
ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 14

ws3.cell(row=1, column=1, value="GEFCom2014-W Task 15 — Probabilistic Forecasting Results").font = TITLE_FONT
ws3.merge_cells("A1:F1")

ws3.cell(row=3, column=1, value="Model trained on 2012 data, tested on Dec 2013 (10 zones x 744 hours)").font = SECTION_FONT
ws3.merge_cells("A3:F3")

gefecom_headers = ["Zone", "Samples", "Model\nPinball Avg", "Benchmark\nPinball Avg",
                   "Model\nP50 MAE", "Benchmark\nP50 MAE"]
for i, h in enumerate(gefecom_headers, 1):
    ws3.cell(row=5, column=i, value=h)
style_header_row(ws3, 5, len(gefecom_headers))

gefecom_data = [
    (1, 737, 0.243087, 0.071143, 0.5532, 0.2034),
    (2, 736, 0.109875, 0.064318, 0.2545, 0.1876),
    (3, 738, 0.121563, 0.077948, 0.2789, 0.2241),
    (4, 738, 0.131539, 0.088118, 0.3012, 0.2557),
    (5, 738, 0.148912, 0.094663, 0.3412, 0.2712),
    (6, 738, 0.142643, 0.095458, 0.3267, 0.2734),
    (7, 738, 0.242435, 0.068957, 0.5621, 0.1987),
    (8, 738, 0.250573, 0.070142, 0.5789, 0.2012),
    (9, 738, 0.187906, 0.068391, 0.4310, 0.1989),
    (10, 738, 0.203835, 0.092774, 0.4701, 0.2667),
    ("OVERALL", 7377, 0.178247, 0.079196, 0.4089, 0.2301),
]

for row_idx, (z, n, mp, bp, mma, bma) in enumerate(gefecom_data, 6):
    vals = [z, n, round(mp, 4), round(bp, 4), round(mma, 4), round(bma, 4)]
    write_row(ws3, row_idx, vals, bold=(z == "OVERALL"))

ws3.cell(row=18, column=1, value="Benchmark = GEFCom2014 persistence model (uses explanatory variables)").font = BODY_FONT
ws3.merge_cells("A18:F18")
ws3.cell(row=19, column=1, value="Pure time-series model (no exp vars) underperforms benchmark by 125% on avg pinball").font = BODY_FONT
ws3.merge_cells("A19:F19")
ws3.cell(row=20, column=1, value="Incorporating exp vars as input features would likely close the gap").font = BODY_FONT
ws3.merge_cells("A20:F20")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: Architecture
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Architecture")
ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 90

ws4.cell(row=1, column=1, value="Model Architecture").font = TITLE_FONT
ws4.merge_cells("A1:B1")

arch_data = [
    ("Component", "Details"),
    ("",
     ""),
    ("1. Input", "x: (batch, seq_len, enc_in) — time series of L steps, N variates"),
    ("",
     "VayuMithra: L=336, N=51 (10 stations x 5 features + ws_s5)"),
    ("",
     "GEFCom2014: L=336, N=50 (10 zones x 5 features: TARGETVAR, U10, V10, U100, V100)"),
    ("",
     ""),
    ("2. Instance Normalization", "Non-stationary Transformer style:"),
    ("",
     "  means = x.mean(1); stdev = sqrt(var(x,1) + 1e-5)"),
    ("",
     "  x = (x - means) / stdev   — per-channel, per-window"),
    ("",
     ""),
    ("3. DataEmbedding_inverted", "(batch, L, N) → permute → (batch, N, L) → Linear(L, d_model) → (batch, N, D)"),
    ("",
     "  Converts each variate's L-length sequence into a D-dimensional token."),
    ("",
     ""),
    ("4. iTransformer Encoder", "Stack of E=2 EncoderLayers, each with:"),
    ("",
     "  a) FullAttention (no causal mask) — (N+time_tokens) x (N+time_tokens)"),
    ("",
     "     Queries, keys, values from LayerNorm((N, D))"),
    ("",
     "  b) Feed-Forward: Linear(D, d_ff) → GELU → Linear(d_ff, D)"),
    ("",
     "  Output: (batch, N, D) — each variate attended to all other variates"),
    ("",
     ""),
    ("5. Per-Variate Projection", "Linear(D, 1): (batch, N, D) → (batch, N, 1) → squeeze → (batch, N)"),
    ("",
     "  Reduces each variate token to a single scalar 'score' for the target."),
    ("",
     "  For multi-target (GEFCom): select TARGETVAR channels → (batch, N_zones)"),
    ("",
     ""),
    ("6a. QuantileProjectionHead", "Linear(N, S * nq): (batch, N) → (batch, S * nq) → view → (batch, S, nq)"),
    ("  (single target)","  Outputs n_quantiles (4 or 99) per horizon step."),
    ("",
     ""),
    ("6b. Per-Target Projection", "Linear(1, S * nq): (batch, N_zones, 1) → (batch, N_zones, S * nq)"),
    ("  (multi-target)","                     → view → (batch, N_zones, S, nq)"),
    ("",
     "  Shares the same projection across all target variates."),
    ("",
     ""),
    ("7. Monotonicity", "Cumulative softplus to enforce P10 ≤ P50 ≤ P90 ≤ P99 (or P1 ≤ ... ≤ P99):"),
    ("",
     "  base = preds[..., :1];  deltas = softplus(preds[..., 1:] - preds[..., :-1])"),
    ("",
     "  preds_ordered = cumsum(concat([base, deltas]), dim=-1)"),
    ("",
     ""),
    ("8. De-normalization", "quantiles = quantiles * stdev_target + means_target"),
    ("",
     "  Transforms quantiles back to original units."),
    ("",
     ""),
    ("Hyperparameters", ""),
    ("  d_model", "128"),
    ("  n_heads", "4"),
    ("  e_layers", "2"),
    ("  d_ff", "512"),
    ("  dropout", "0.15"),
    ("  activation", "GELU"),
    ("  lr", "3e-4 (AdamW)"),
    ("  weight_decay", "1e-3"),
    ("  batch_size", "128 (VayuMithra) / 64 (GEFCom)"),
    ("  scheduler", "CosineAnnealingLR(T_max=30..50)"),
    ("  total params", "~440K"),
]

for i, (k, v) in enumerate(arch_data, 3):
    c1 = ws4.cell(row=i, column=1, value=k)
    c2 = ws4.cell(row=i, column=2, value=v)
    c1.font = BOLD if k and not k.startswith("  ") and k != "Component" else BODY_FONT
    c2.font = BODY_FONT
    c1.border = BORDER
    c2.border = BORDER

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5: Loss_Functions
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Loss_Functions")
ws5.column_dimensions["A"].width = 30
ws5.column_dimensions["B"].width = 90

ws5.cell(row=1, column=1, value="Loss Functions & Evaluation Metrics").font = TITLE_FONT
ws5.merge_cells("A1:B1")

loss_data = [
    ("Metric", "Definition"),
    ("",
     ""),
    ("1. Pinball Loss", "Primary training loss for quantile regression."),
    ("  (Quantile Loss)", ""),
    ("", "L_q(y, y_hat) = max(q * (y - y_hat), (q-1) * (y - y_hat))"),
    ("",
     "  = q * (y - y_hat)   if y >= y_hat, else (q-1) * (y - y_hat)"),
    ("",
     "  where q ∈ (0,1) is the quantile level, y is the true value,"),
    ("",
     "  and y_hat is the predicted quantile."),
    ("",
     "  Properties:"),
    ("", "  - Asymmetric: penalizes over/under-prediction differently"),
    ("", "  - Proper scoring rule: minimized when y_hat = true q-quantile"),
    ("", "  - L_0.5 = 0.5 * |y - y_hat| = half of MAE"),
    ("", ""),
    ("2. Multi-Quantile Pinball", "Average pinball across all quantile levels:"),
    ("", "  Loss = (1/K) * SUM_k L_{q_k}(y, y_hat_k)"),
    ("", "  where K = number of quantiles (4 for VayuMithra, 99 for GEFCom)."),
    ("", ""),
    ("3. Quantile Monotonicity", "Cumulative softplus ensures P10 ≤ P50 ≤ P90 ≤ P99:"),
    ("", "  Let r_1, r_2, ..., r_K be raw outputs from the projection head."),
    ("", "  y_1 = r_1"),
    ("", "  y_k = y_{k-1} + softplus(r_k - y_{k-1})   for k = 2..K"),
    ("", "  This guarantees y_1 ≤ y_2 ≤ ... ≤ y_K without constraints."),
    ("", ""),
    ("4. Interval Coverage", "Fraction of true values falling inside the prediction interval:"),
    ("", "  Coverage = (1/N) * SUM_i 1{lo_i ≤ y_i ≤ hi_i}"),
    ("", "  Target: ~80% for P10-P90 interval (i.e., q=0.1 to q=0.9)."),
    ("", ""),
    ("5. Interval Width", "Mean width of the prediction interval:"),
    ("", "  Width = (1/N) * SUM_i (hi_i - lo_i)"),
    ("", "  Measured in original units (m/s for wind speed, fraction for power)."),
    ("", ""),
    ("6. Winkler Score", "Combines width + penalties for misses (sharpness + calibration):"),
    ("", "  Winkler = (1/N) * SUM_i [ (hi_i - lo_i)" ),
    ("", "    + (2/alpha) * max(lo_i - y_i, 0)"),
    ("", "    + (2/alpha) * max(y_i - hi_i, 0) ]"),
    ("", "  where alpha = 0.20 for P10-P90 interval."),
    ("", "  Lower is better. Punishes both wide intervals and missed targets."),
    ("", ""),
    ("7. MAE (P50)", "Mean Absolute Error of the median (P50) prediction:"),
    ("", "  MAE = (1/N) * SUM_i |y_i - y_hat_i_P50|"),
    ("", ""),
    ("8. RMSE (P50)", "Root Mean Squared Error of the median:"),
    ("", "  RMSE = sqrt((1/N) * SUM_i (y_i - y_hat_i_P50)^2)"),
]

for i, (k, v) in enumerate(loss_data, 3):
    c1 = ws5.cell(row=i, column=1, value=k)
    c2 = ws5.cell(row=i, column=2, value=v)
    c1.font = BOLD if k and not k.startswith("  ") and k != "Metric" else BODY_FONT
    c2.font = BODY_FONT
    c1.border = BORDER
    c2.border = BORDER

# ── Save ────────────────────────────────────────────────────────────────────
out_path = r"C:\Projects\raghavan\vayumithra_research\results\probabilistic_forecasting_report.xlsx"
wb.save(out_path)
print(f"Report saved to {out_path}", flush=True)
